"""Security: report exporters must HTML-escape PCE-derived values (stored XSS)
and neutralize xlsx cell values against spreadsheet/CSV formula injection.

Each test feeds an attacker-style payload as a PCE-derived field and asserts the
raw payload does NOT survive into the rendered output (it appears escaped, e.g.
``&lt;img``; for xlsx the cell is stored as text, not a live formula).
"""
from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from src.report.exporters.report_i18n import STRINGS

PAYLOAD = '<img src=x onerror=alert(1)>'
ESCAPED_MARK = '&lt;img'  # html.escape turns '<' into '&lt;'


def _en_s():
    return lambda k: STRINGS[k].get("en") or STRINGS[k]["en"]


# ── Traffic / SecurityRisk (html_exporter.py) ────────────────────────────────

def test_traffic_table_cell_is_escaped():
    from src.report.exporters.html_exporter import _df_to_html
    df = pd.DataFrame([{"Hostname": PAYLOAD, "Process": PAYLOAD}])
    html = _df_to_html(df)
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_traffic_format_evidence_is_escaped():
    from src.report.exporters.html_exporter import _format_evidence
    html = _format_evidence({"hostname": PAYLOAD}, lang="en")
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_traffic_findings_html_escapes_pce_values():
    from src.report.exporters.html_exporter import SecurityRiskHtmlExporter
    from src.report.rules._base import Finding
    finding = Finding(
        rule_id="ZZ99",  # absent from STRINGS -> rule_name falls back to f.rule_name
        rule_name=PAYLOAD,
        severity="HIGH",
        category="LateralMovement",
        description=PAYLOAD,
        recommendation=PAYLOAD,
        evidence={"hostname": PAYLOAD},
    )
    exporter = SecurityRiskHtmlExporter({"findings": [finding]}, lang="en")
    exporter._s = _en_s()
    html = exporter._findings_html()
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_traffic_attack_summary_html_is_escaped():
    from src.report.exporters.html_exporter import SecurityRiskHtmlExporter
    exporter = SecurityRiskHtmlExporter({}, lang="en")
    exporter._s = _en_s()
    mod12 = {
        "boundary_breaches": [{"severity": "HIGH", "finding": PAYLOAD, "action": PAYLOAD}],
        "action_matrix": [{"action_code": PAYLOAD, "action": PAYLOAD}],
    }
    html = exporter._attack_summary_html(mod12)
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_traffic_ringfence_html_is_escaped():
    from src.report.exporters.html_exporter import SecurityRiskHtmlExporter
    exporter = SecurityRiskHtmlExporter(
        {"mod_ringfence": {"top_apps": [{"app": PAYLOAD, "flows": 5}]}}, lang="en")
    exporter._s = _en_s()
    html = exporter._mod_ringfence_html()
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_traffic_key_findings_html_is_escaped():
    # key_findings_html is built inline inside _build(); exercise the full render.
    from src.report.exporters.html_exporter import SecurityRiskHtmlExporter
    results = {
        "findings": [],
        "mod12": {"key_findings": [{"severity": "HIGH", "finding": PAYLOAD, "action": PAYLOAD}]},
    }
    exporter = SecurityRiskHtmlExporter(results, lang="en")
    html = exporter._build()
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


# ── Audit (audit_html_exporter.py) ───────────────────────────────────────────

def test_audit_table_cell_is_escaped():
    from src.report.exporters.audit_html_exporter import _df_to_html
    df = pd.DataFrame([{"actor": PAYLOAD, "resource_name": PAYLOAD}])
    html = _df_to_html(df, lang="en")
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_audit_high_impact_provisions_is_escaped():
    from src.report.exporters.audit_html_exporter import AuditHtmlExporter
    exporter = AuditHtmlExporter({}, lang="en")
    exporter._s = _en_s()
    items = [{
        "workloads_affected": 99,
        "timestamp": PAYLOAD,
        "event_type": PAYLOAD,
        "actor": PAYLOAD,
        "src_ip": PAYLOAD,
        "resource_name": PAYLOAD,
        "status": PAYLOAD,
    }]
    html = exporter._high_impact_provisions_html(items, threshold=50)
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


# ── VEN (ven_html_exporter.py) ───────────────────────────────────────────────

def test_ven_table_cell_is_escaped():
    from src.report.exporters.ven_html_exporter import VenHtmlExporter
    df = pd.DataFrame([{"hostname": PAYLOAD, "os": PAYLOAD}])
    exporter = VenHtmlExporter({"online": df}, lang="en")
    html = exporter._build()
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


# ── Policy Usage (policy_usage_html_exporter.py) ─────────────────────────────

def test_policy_usage_table_cell_is_escaped():
    from src.report.exporters.policy_usage_html_exporter import _df_to_html
    df = pd.DataFrame([{"Ruleset": PAYLOAD, "Source": PAYLOAD}])
    html = _df_to_html(df, lang="en")
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_policy_usage_attention_html_is_escaped():
    from src.report.exporters.policy_usage_html_exporter import PolicyUsageHtmlExporter
    exporter = PolicyUsageHtmlExporter({}, lang="en")
    exporter._s = _en_s()
    html = exporter._attention_html([{"ruleset": PAYLOAD, "unused_count": 3}])
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


def test_policy_usage_execution_notes_is_escaped():
    from src.report.exporters.policy_usage_html_exporter import PolicyUsageHtmlExporter
    exporter = PolicyUsageHtmlExporter({}, lang="en")
    exporter._s = _en_s()
    html = exporter._execution_html({"execution_notes": [PAYLOAD]})
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


# ── Concern cards (concern_card.py) ──────────────────────────────────────────

def test_concern_cards_are_escaped():
    from src.report.exporters.concern_card import render_concern_cards
    items = [{
        "risk": "HIGH",
        "event_type": PAYLOAD,
        "count": 2,
        "summary": PAYLOAD,
        "recommendation": PAYLOAD,
        "actors": [PAYLOAD],
        "targets": [PAYLOAD],
        "resources": [PAYLOAD],
        "src_ips": [PAYLOAD],
    }]
    html = render_concern_cards(items, lang="en")
    assert PAYLOAD not in html
    assert ESCAPED_MARK in html


# ── XLSX formula / CSV injection (xlsx_exporter.py) ──────────────────────────

def test_xlsx_neutralizes_formula_injection(tmp_path):
    from src.report.exporters.xlsx_exporter import export_xlsx
    formula = '=HYPERLINK("http://attacker","pwn")'
    result = {
        "record_count": 1,
        "metadata": {"title": "t"},
        "module_results": {
            "mod": {
                "summary": "",
                "table": [{"Hostname": formula, "=Header": "x", "Note": "+1+2"}],
            }
        },
    }
    out = tmp_path / "f.xlsx"
    export_xlsx(result, str(out))
    wb = load_workbook(str(out))
    ws = wb["mod"]
    # No cell may be stored as a live formula.
    for row in ws.iter_rows():
        for cell in row:
            assert cell.data_type != "f", f"formula injection at {cell.coordinate}: {cell.value!r}"
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    # The dangerous header and value are quote-neutralized, not stored verbatim.
    assert formula not in flat
    assert any(str(v).startswith("'=HYPERLINK") for v in flat)
    assert any(str(v).startswith("'=Header") for v in flat)
