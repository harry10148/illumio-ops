"""Audit XLSX 建構器只讀 module_results（不重算），必須產出真實分頁與資料列。"""
import pandas as pd
import pytest
from openpyxl import load_workbook
from unittest.mock import MagicMock


@pytest.fixture
def sample_module_results():
    """手組 module_results dict，模擬 _run_pipeline 產出的各模組結果。"""
    return {
        "mod00": {
            "kpis": [
                {"label": "Total Events", "value": 42},
                {"label": "Distinct Actors", "value": 3},
            ],
        },
        "mod01": {
            "summary": pd.DataFrame([
                {"metric": "success_rate", "value": "98%"},
                {"metric": "failure_count", "value": 2},
            ]),
        },
        "mod02": {
            "per_user": pd.DataFrame([
                {"actor": "admin", "event_count": 10},
            ]),
        },
        "mod03": {
            "summary": pd.DataFrame([
                {"event_type": "policy.create", "count": 1},
            ]),
        },
        "mod04": {
            "correlated_sequences": pd.DataFrame([
                {"sequence_id": "seq-corr-001", "actor": "admin"},
            ]),
            "brute_force_detections": pd.DataFrame([
                {"actor": "svc-deploy", "attempts": 5},
            ]),
            "off_hours_operations": pd.DataFrame([
                {"actor": "admin", "hour": 3},
            ]),
        },
    }


def test_audit_xlsx_has_expected_sheets(sample_module_results, tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Attention Required", "Health", "Users", "Policy Changes", "Correlations"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"


def test_audit_xlsx_policy_changes_has_rows(sample_module_results, tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    rows = list(wb["Policy Changes"].iter_rows(values_only=True))
    assert len(rows) >= 2, "Policy Changes sheet should have header + data row"
    flat = [v for r in rows for v in r]
    assert "policy.create" in flat


def test_audit_xlsx_correlations_has_correlated_sequences_data(sample_module_results, tmp_path):
    """RED 關鍵斷言：舊路徑只寫純量 KPI，correlated_sequences 這種 DataFrame 恆被跳過而永遠空白。"""
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    flat = [v for r in wb["Correlations"].iter_rows(values_only=True) for v in r]
    assert "seq-corr-001" in flat, "Correlations sheet must contain correlated_sequences row data"


def test_audit_xlsx_attention_has_kpi_label(sample_module_results, tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    flat = [v for r in wb["Attention Required"].iter_rows(values_only=True) for v in r]
    assert "Total Events" in flat


def test_audit_xlsx_missing_keys_write_empty_note_without_crash(tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx({}, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Attention Required", "Health", "Users", "Policy Changes", "Correlations"}
    assert expected.issubset(set(wb.sheetnames))


def test_audit_generator_export_xlsx_writes_real_rows(sample_module_results, tmp_path):
    """整合測試：export(fmt='xlsx') 走完整路徑，openpyxl 讀回應有非空分頁。"""
    from src.report.audit_generator import AuditGenerator, AuditReportResult

    cm = MagicMock()
    cm.config = {
        "api": {"url": "https://pce.test", "org_id": "1", "key": "k", "secret": "s",
                "verify_ssl": False},
    }
    gen = AuditGenerator(cm, api_client=MagicMock())
    result = AuditReportResult(
        record_count=3,
        date_range=("2026-04-25", "2026-04-25"),
        module_results=sample_module_results,
        dataframe=pd.DataFrame(),
    )
    paths = gen.export(result, fmt="xlsx", output_dir=str(tmp_path))
    xlsx_paths = [p for p in paths if p.endswith(".xlsx")]
    assert xlsx_paths, "export(fmt='xlsx') should return an xlsx path"

    wb = load_workbook(xlsx_paths[0])
    non_empty_sheets = [
        name for name in wb.sheetnames
        if any(any(v is not None for v in row) for row in wb[name].iter_rows(values_only=True))
    ]
    assert non_empty_sheets, "at least one sheet must have data rows"
