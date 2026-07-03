"""VEN XLSX：generate_ven_xlsx 只讀 _analyze() 回傳的 analysis dict（不得自行分桶）。

心跳邊界案例鎖住 1 小時上線門檻（spec K1）：
  - 59 分鐘前心跳 → online
  - 61 分鐘前心跳 → offline，且落在 lost_today（<24h）
  - 30 小時前心跳 → offline，落在 lost_yesterday（24-48h）
  - 72 小時前心跳 → offline，兩個 lost 桶都不落
"""
import types
from datetime import datetime, timedelta, timezone

import pytest
from openpyxl import load_workbook

from src.report.ven_status_generator import VenStatusGenerator


def _workload(host, minutes_ago=None, hours_ago=None, ven_status="active", version="21.5.35"):
    now = datetime.now(timezone.utc)
    if hours_ago is not None:
        hb = now - timedelta(hours=hours_ago)
    else:
        hb = now - timedelta(minutes=minutes_ago)
    return {
        "hostname": host,
        "interfaces": [{"address": "10.0.0.5"}],
        "labels": [],
        "agent": {"status": {
            "status": ven_status,
            # hours_since_last_heartbeat 刻意不提供，逼 _analyze 走 last_heartbeat_on 時間戳分支
            "security_policy_sync_state": "synced",
            "last_heartbeat_on": hb.isoformat(),
            "security_policy_refresh_at": hb.isoformat(),
            "managed_since": "2024-01-01T00:00:00Z",
            "agent_version": version,
        }},
    }


@pytest.fixture
def analysis():
    """走真實 VenStatusGenerator._analyze()，取得心跳邊界案例的分桶結果。"""
    cm = types.SimpleNamespace(config={"settings": {"timezone": "UTC"}})
    g = VenStatusGenerator(cm, api_client=object())
    g._lang = "en"
    workloads = [
        _workload("web-59m", minutes_ago=59, version="21.5.35"),
        _workload("web-61m", minutes_ago=61, version="21.5.35"),
        _workload("web-30h", hours_ago=30, version="21.5.40"),
        _workload("web-72h", hours_ago=72, version="21.5.40"),
    ]
    df = g._build_dataframe(workloads)
    return g._analyze(df)


def _hostnames(df):
    if df is None or df.empty:
        return set()
    return set(df["Hostname"].tolist())


def test_heartbeat_boundary_buckets_are_correct(analysis):
    """鎖定 _analyze 的 1 小時心跳分桶語意（59m→online、61m→lost_today）。

    註：此測試直接驗 _analyze（該邏輯本就正確、非本次修改標的），
    作為 K1 語意的回歸錨點；舊 generate_ven_xlsx 的 24h 門檻缺陷由
    test_ven_xlsx_builder_no_self_bucketing_uses_provided_dict 與
    sheets_match_analysis_buckets 從「禁止自行分桶」面向鎖住。
    """
    assert _hostnames(analysis["online"]) == {"web-59m"}
    assert _hostnames(analysis["offline"]) == {"web-61m", "web-30h", "web-72h"}
    assert _hostnames(analysis["lost_today"]) == {"web-61m"}
    assert _hostnames(analysis["lost_yesterday"]) == {"web-30h"}


def test_ven_xlsx_sheets_match_analysis_buckets(analysis, tmp_path):
    from src.report.ven_status_generator import generate_ven_xlsx

    out_path = tmp_path / "ven.xlsx"
    generate_ven_xlsx(analysis, str(out_path))
    wb = load_workbook(str(out_path))

    expected_sheets = {"Online", "Offline", "Lost Today", "Lost Yesterday", "Versions"}
    assert expected_sheets.issubset(set(wb.sheetnames)), (
        f"missing: {expected_sheets - set(wb.sheetnames)}"
    )

    def _sheet_hostnames(name):
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            return set()
        header = rows[0]
        if "Hostname" not in header:
            return set()
        idx = header.index("Hostname")
        return {r[idx] for r in rows[1:] if r[idx] is not None}

    assert _sheet_hostnames("Online") == _hostnames(analysis["online"])
    assert _sheet_hostnames("Offline") == _hostnames(analysis["offline"])
    assert _sheet_hostnames("Lost Today") == _hostnames(analysis["lost_today"])
    assert _sheet_hostnames("Lost Yesterday") == _hostnames(analysis["lost_yesterday"])


def test_ven_xlsx_versions_sheet_matches_by_version(analysis, tmp_path):
    from src.report.ven_status_generator import generate_ven_xlsx

    out_path = tmp_path / "ven.xlsx"
    generate_ven_xlsx(analysis, str(out_path))
    wb = load_workbook(str(out_path))
    rows = list(wb["Versions"].iter_rows(values_only=True))
    flat_pairs = {(r[0], r[1]) for r in rows[1:]}
    expected_pairs = set(analysis["by_version"].items())
    assert flat_pairs == expected_pairs


def test_ven_xlsx_builder_no_self_bucketing_uses_provided_dict(tmp_path):
    """generate_ven_xlsx 不得自行分桶：傳入手組 analysis dict 時，輸出必須原樣反映該 dict。"""
    import pandas as pd
    from src.report.ven_status_generator import generate_ven_xlsx

    fake_analysis = {
        "online": pd.DataFrame([{"Hostname": "only-online-host"}]),
        "offline": pd.DataFrame([{"Hostname": "only-offline-host"}]),
        "lost_today": pd.DataFrame(columns=["Hostname"]),
        "lost_yesterday": pd.DataFrame(columns=["Hostname"]),
        "by_version": {"9.9.9": 1},
    }
    out_path = tmp_path / "ven.xlsx"
    generate_ven_xlsx(fake_analysis, str(out_path))
    wb = load_workbook(str(out_path))

    online_rows = [v for r in wb["Online"].iter_rows(values_only=True) for v in r]
    offline_rows = [v for r in wb["Offline"].iter_rows(values_only=True) for v in r]
    assert "only-online-host" in online_rows
    assert "only-offline-host" not in online_rows
    assert "only-offline-host" in offline_rows


def test_ven_status_generator_export_xlsx_writes_real_rows(analysis, tmp_path):
    """整合測試：export(fmt='xlsx') 走完整路徑，openpyxl 讀回應有非空分頁。"""
    from unittest.mock import MagicMock
    from src.report.ven_status_generator import VenStatusGenerator, VenStatusResult

    cm = MagicMock()
    gen = VenStatusGenerator(cm, api_client=MagicMock())
    result = VenStatusResult(
        record_count=4,
        module_results=analysis,
        dataframe=None,
    )
    paths = gen.export(result, fmt="xlsx", output_dir=str(tmp_path))
    xlsx_paths = [p for p in paths if p.endswith(".xlsx")]
    assert xlsx_paths, "export(fmt='xlsx') should return an xlsx path"

    wb = load_workbook(xlsx_paths[0])
    non_empty_sheets = [
        name for name in wb.sheetnames
        if any(any(v is not None for v in row) for row in wb[name].iter_rows(values_only=True))
    ]
    assert non_empty_sheets, "at least one sheet must have data rows"

    online_flat = [v for r in wb["Online"].iter_rows(values_only=True) for v in r]
    assert "web-59m" in online_flat
