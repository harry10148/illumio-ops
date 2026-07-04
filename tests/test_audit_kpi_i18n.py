"""Audit KPI 標籤中文化（spec I2）。

_resolve_label 解析順序：label_key/i18n_key 可解析（回值 != key 名，且非
[MISSING:...]）時優先；否則 fallback 到 label；再否則 fallback 到 key 本身。
涵蓋 exec summary 共用面（audit/mod12/ven/pu 四種 kpis 形狀）以及 audit XLSX
Attention sheet 的同步。
"""
from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from src.report.exporters._exec_summary import _resolve_label, render_exec_summary_html


def test_audit_kpi_resolves_zh():
    # audit_mod00_executive.py 形狀：label 為固定英文常數，label_key 才是真正
    # 可在地化的來源。
    kpi = {"label_key": "rpt_au_kpi_total_events", "label": "Total Events", "value": "12"}
    assert _resolve_label(kpi, lang="zh_TW") == "事件總數"


def test_fallback_to_label_when_key_missing():
    kpi = {"label_key": "rpt_no_such_key_xyz", "label": "Fallback", "value": "1"}
    assert _resolve_label(kpi, lang="zh_TW") == "Fallback"


def test_mod12_style_kpi_unchanged():
    # mod12_executive_summary.py 形狀：label 產生時已經用 t(label_key, lang)
    # 解析過，label_key 同值可解析——新舊邏輯輸出應一致。
    kpi = {"label_key": "mod12_kpi_total_flows", "label": "流量總數", "value": "9"}
    assert _resolve_label(kpi, lang="zh_TW") == "流量總數"


def test_ven_style_kpi_i18n_key_resolves():
    # ven_status_generator.py 形狀：只帶 lang-agnostic 'i18n_key'，沒有 label。
    kpi = {"i18n_key": "rpt_ven_kpi_total", "value": "5"}
    assert _resolve_label(kpi, lang="zh_TW") == "VEN 總數"
    assert _resolve_label(kpi, lang="en") == "Total VENs"


def test_pu_style_kpi_label_key_resolves():
    # policy_usage pu_mod00_executive.py 形狀：label 已由生成時的 t() 在地化，
    # label_key 同值可解析——輸出應與預先解析的 label 相同。
    kpi = {
        "label_key": "rpt_pu_total_rules",
        "label": "啟用中的規則總數",
        "value": "10",
    }
    assert _resolve_label(kpi, lang="zh_TW") == "啟用中的規則總數"


def test_render_exec_summary_html_zh_shows_chinese_label():
    """渲染級回歸：audit 風格 kpi 在 zh_TW 下 exec summary 真的顯示中文標籤。"""
    mod00 = {
        "kpis": [
            {"label_key": "rpt_au_kpi_total_events", "label": "Total Events", "value": "42"},
        ],
    }
    html = render_exec_summary_html(mod00, report_name="稽核報表", lang="zh_TW")
    assert "事件總數" in html
    assert "Total Events" not in html


def test_render_exec_summary_html_en_still_english():
    mod00 = {
        "kpis": [
            {"label_key": "rpt_au_kpi_total_events", "label": "Total Events", "value": "42"},
        ],
    }
    html = render_exec_summary_html(mod00, report_name="Audit Report", lang="en")
    assert "Total Events" in html


def test_audit_xlsx_attention_sheet_shows_chinese_label(tmp_path):
    """XLSX Attention sheet 同步：label_key 在 lang=zh_TW 時優先解析，不再固定英文。"""
    from src.report.audit_generator import generate_audit_xlsx

    module_results = {
        "mod00": {
            "kpis": [
                {"label_key": "rpt_au_kpi_total_events", "label": "Total Events", "value": 42},
            ],
        },
    }
    out_path = tmp_path / "audit_zh.xlsx"
    generate_audit_xlsx(module_results, str(out_path), lang="zh_TW")
    wb = load_workbook(str(out_path))
    flat = [v for r in wb["待處理事項"].iter_rows(values_only=True) for v in r]
    assert "事件總數" in flat
    assert "Total Events" not in flat


def test_audit_html_exporter_no_dead_kpi_cards_reference():
    """順手清理回歸：_kpi_label/kpi_cards 死碼已移除，_build() 仍正常輸出。"""
    from src.report.exporters.audit_html_exporter import AuditHtmlExporter

    results = {
        "mod00": {
            "kpis": [{"label_key": "rpt_au_kpi_total_events", "label": "Total Events", "value": "42"}],
            "generated_at": "2026-05-15 09:00",
            "attention_items": [],
            "severity_dist": {},
            "top_events_overall": pd.DataFrame(),
        },
        "mod01": {}, "mod02": {}, "findings": [],
    }
    html = AuditHtmlExporter(results, pd.DataFrame())._build()
    # The shared stylesheet still defines .kpi-card (used by other exporters),
    # but AuditHtmlExporter must not emit any <div class="kpi-card"> markup.
    assert '<div class="kpi-card">' not in html
