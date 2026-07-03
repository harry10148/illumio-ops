"""xlsx_exporter 共用 DataFrame 寫入層。"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from src.report.exporters.xlsx_exporter import add_df_sheet, add_stacked_tables_sheet


def _read_rows(path, sheet):
    wb = load_workbook(path)
    return [[c.value for c in r] for r in wb[sheet].iter_rows()]


def test_add_df_sheet_writes_header_and_rows(tmp_path):
    wb = Workbook()
    df = pd.DataFrame([{"Name": "web", "Count": 3}, {"Name": "db", "Count": 1}])
    add_df_sheet(wb, "MySheet", df)
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    rows = _read_rows(p, "MySheet")
    assert rows[0] == ["Name", "Count"]
    assert rows[1] == ["web", 3]
    assert len(rows) == 3


def test_formula_injection_neutralized(tmp_path):
    wb = Workbook()
    df = pd.DataFrame([{"Name": "=HYPERLINK(evil)", "Count": 1}])
    add_df_sheet(wb, "Inj", df)
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    rows = _read_rows(p, "Inj")
    assert rows[1][0].startswith("'=")  # 前綴單引號中和


def test_empty_df_writes_note(tmp_path):
    wb = Workbook()
    add_df_sheet(wb, "Empty", pd.DataFrame(), empty_note="nothing here")
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    rows = _read_rows(p, "Empty")
    assert "nothing here" in [v for r in rows for v in r]


def test_stacked_tables_have_subtitles(tmp_path):
    wb = Workbook()
    t1 = pd.DataFrame([{"A": 1}])
    t2 = pd.DataFrame([{"B": 2}])
    add_stacked_tables_sheet(wb, "Stack", [("First", t1), ("Second", t2)])
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    flat = [v for r in _read_rows(p, "Stack") for v in r]
    assert "First" in flat and "Second" in flat and 1 in flat and 2 in flat
