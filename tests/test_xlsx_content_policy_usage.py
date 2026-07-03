"""Policy Usage XLSX 建構器只讀 module_results（不重算），必須產出真實分頁與資料列。"""
import pandas as pd
import pytest
from openpyxl import load_workbook
from unittest.mock import MagicMock


@pytest.fixture
def sample_module_results():
    """手組 module_results dict，模擬 _run_pipeline 產出的各模組結果。"""
    unused_df = pd.DataFrame([
        {"Ruleset": "default", "No": 3, "Rule ID": "r-3", "Type": "Allow",
         "Description": "old broad allow", "Destination": "Any", "Source": "Any",
         "Services": "All Services", "Observed Hit Ports": "None in lookback",
         "Enabled": True, "Created At": "2026-01-01"},
        {"Ruleset": "default", "No": 4, "Rule ID": "r-4", "Type": "Deny",
         "Description": "deny untrusted", "Destination": "Any", "Source": "Any",
         "Services": "All Services", "Observed Hit Ports": "None in lookback",
         "Enabled": True, "Created At": "2026-01-02"},
    ])
    return {
        "mod01": {
            "total_rules": 4,
            "hit_count": 2,
            "unused_count": 2,
            "hit_rate_pct": 50.0,
            "summary_df": pd.DataFrame([
                {"Status": "Hit", "Count": 2, "Percentage": "50.0%"},
                {"Status": "Unused", "Count": 2, "Percentage": "50.0%"},
            ]),
        },
        "mod02": {
            "hit_df": pd.DataFrame([
                {"Ruleset": "default", "No": 1, "Rule ID": "r-1", "Type": "Allow",
                 "Description": "allow web to db", "Destination": "db", "Source": "web",
                 "Services": "TCP/443", "Hit Count": 1234, "Top Hit Ports": "443/tcp (1234)",
                 "Enabled": True},
            ]),
            "top_ports_df": pd.DataFrame([{"Port / Proto": "443/tcp", "Flow Count": 1234}]),
            "chart_spec": None,
        },
        "mod03": {
            "unused_df": unused_df,
        },
        "mod04": {
            "total_deny": 2,
            "total_allow": 2,
            "deny_ratio_pct": 50.0,
            "deny_hit_count": 1,
            "deny_unused_count": 1,
            "deny_hit_rate_pct": 50.0,
            "override_deny_count": 0,
            "deny_detail_df": pd.DataFrame([
                {"Ruleset": "default", "Rule No": 2, "Type": "Deny",
                 "Description": "deny untrusted", "Scope": "Broad", "Blocked Flows": 5,
                 "Status": "Hit", "Targets High-Risk Ports": False, "Source": "Any",
                 "Destination": "Any", "Services": "All Services"},
            ]),
            "deny_summary_df": pd.DataFrame([
                {"Category": "Deny Rules (Hit)", "Count": 1, "Pct": "50.0%"},
                {"Category": "Deny Rules (Unused)", "Count": 1, "Pct": "50.0%"},
                {"Category": "Override Deny", "Count": 0, "Pct": "0.0%"},
            ]),
        },
    }


def test_policy_usage_xlsx_has_expected_sheets(sample_module_results, tmp_path):
    from src.report.policy_usage_generator import generate_policy_usage_xlsx
    out_path = tmp_path / "pu.xlsx"
    generate_policy_usage_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Hit Rules", "Unused Rules", "Deny Effectiveness", "Execution Stats"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"


def test_policy_usage_xlsx_unused_sheet_rowcount_matches_source(sample_module_results, tmp_path):
    """同源驗證：Unused sheet 資料列數必須等於 fixture unused_df 列數（無重算截斷）。"""
    from src.report.policy_usage_generator import generate_policy_usage_xlsx
    out_path = tmp_path / "pu.xlsx"
    generate_policy_usage_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    sheet = wb["Unused Rules"]
    rows = list(sheet.iter_rows(values_only=True))
    expected_rows = len(sample_module_results["mod03"]["unused_df"])
    assert len(rows) - 1 == expected_rows, (
        f"expected {expected_rows} data rows (+1 header), got {len(rows)} total rows"
    )


def test_policy_usage_xlsx_deny_sheet_has_summary_and_detail(sample_module_results, tmp_path):
    from src.report.policy_usage_generator import generate_policy_usage_xlsx
    out_path = tmp_path / "pu.xlsx"
    generate_policy_usage_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    flat = [v for r in wb["Deny Effectiveness"].iter_rows(values_only=True) for v in r]
    assert "Deny Rules (Hit)" in flat, "Deny Effectiveness must contain deny_summary_df data"
    assert "deny untrusted" in flat, "Deny Effectiveness must contain deny_detail_df data"


def test_policy_usage_xlsx_execution_stats_has_total_rules(sample_module_results, tmp_path):
    from src.report.policy_usage_generator import generate_policy_usage_xlsx
    out_path = tmp_path / "pu.xlsx"
    generate_policy_usage_xlsx(sample_module_results, str(out_path))
    wb = load_workbook(str(out_path))
    rows = list(wb["Execution Stats"].iter_rows(values_only=True))
    flat = [v for r in rows for v in r]
    assert "Total Rules" in flat
    idx = flat.index("Total Rules")
    assert flat[idx + 1] == sample_module_results["mod01"]["total_rules"]


def test_policy_usage_xlsx_missing_keys_write_empty_note_without_crash(tmp_path):
    from src.report.policy_usage_generator import generate_policy_usage_xlsx
    out_path = tmp_path / "pu.xlsx"
    generate_policy_usage_xlsx({}, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Hit Rules", "Unused Rules", "Deny Effectiveness", "Execution Stats"}
    assert expected.issubset(set(wb.sheetnames))


def test_policy_usage_generator_export_xlsx_writes_real_rows(sample_module_results, tmp_path):
    """整合測試：export(fmt='xlsx') 走完整路徑，openpyxl 讀回應有非空分頁。"""
    from src.report.policy_usage_generator import PolicyUsageGenerator, PolicyUsageResult

    cm = MagicMock()
    gen = PolicyUsageGenerator(cm, api_client=MagicMock())
    result = PolicyUsageResult(
        record_count=4,
        date_range=("2026-04-25", "2026-04-25"),
        module_results=sample_module_results,
        dataframe=pd.DataFrame(),
    )
    paths = gen.export(result, fmt="xlsx", output_dir=str(tmp_path))
    xlsx_paths = [p for p in paths if p.endswith(".xlsx")]
    assert xlsx_paths, "export(fmt='xlsx') should return an xlsx path"

    wb = load_workbook(xlsx_paths[0])
    non_empty_sheets = [
        name for name in wb.sheetnames
        if any(any(v is not None for v in row) for row in wb[name].iter_rows(values_only=True))
    ]
    assert non_empty_sheets, "at least one sheet must have data rows"
