"""Traffic curated workbook 建構器只讀 module_results（不重算），必須產出真實分頁與資料列。

build_traffic_xlsx 對應 _run_pipeline 的 module_results：
  mod12→Executive Summary、mod02→Policy Decisions、mod03→Uncovered Flows、
  mod15→Lateral Movement（含六下放表）、mod07→Cross-Label Matrix（四維度）、
  mod08→Unmanaged Hosts（含三下放表）。缺 key（該 profile 不跑該模組）→ 整張 sheet 略過。
"""
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def full_module_results():
    """手組六模組齊備的 module_results dict（security_risk profile 全跑）。"""
    return {
        "mod12": {
            "kpis": [
                {"label": "Total Flows", "value": "7"},
                {"label": "Blocked Flows", "value": "1"},
            ],
            "chart_spec": None,
        },
        "mod02": {
            "summary": pd.DataFrame([{"Decision": "allowed", "Flows": 5}]),
            "port_coverage": pd.DataFrame([{"Port": 445, "Covered": "yes"}]),
            "audit_flags": pd.DataFrame([{"Src": "10.0.0.1", "Flag": "review"}]),
            "chart_spec": None,
        },
        "mod03": {
            "top_flows": pd.DataFrame([{"Source": "10.0.0.4", "Destination": "10.0.0.2", "Port": 22}]),
            "uncovered_port_services": pd.DataFrame([{"Port": 22, "Service": "ssh"}]),
            "uncovered_ports": pd.DataFrame([{"Port": 22, "Count": 1}]),
            "uncovered_services": pd.DataFrame([{"Service": "ssh", "Count": 1}]),
            "by_recommendation": pd.DataFrame([{"Recommendation": "intra_app", "Count": 1}]),
        },
        "mod15": {
            "service_summary": pd.DataFrame([{"Service": "SMB", "Count": 5}]),
            "fan_out_sources": pd.DataFrame([{"Src": "10.0.0.1", "Fanout": 3}]),
            "allowed_lateral_flows": pd.DataFrame([{"Src": "10.0.0.1", "Dst": "10.0.0.2"}]),
            "attack_paths": pd.DataFrame([{"Path": "a->b->c"}]),
            "ip_top_talkers": pd.DataFrame([{"Host": "10.0.0.1", "Connections": 10}]),
            "ip_top_pairs": pd.DataFrame([{"Src": "10.0.0.1", "Dst": "10.0.0.2", "Count": 5}]),
            "source_risk_scores": pd.DataFrame([{"Source": "10.0.0.1", "Risk": 0.8}]),
            "bridge_nodes": pd.DataFrame([{"Node": "10.0.0.1", "Score": 0.6}]),
            "top_reachable_nodes": pd.DataFrame([{"Node": "10.0.0.2", "Depth": 3}]),
            "app_chains": pd.DataFrame([{"Chain": "app1->app2", "Count": 2}]),
            "chart_spec": None,
        },
        "mod07": {
            "matrices": {
                "env": {"top_cross_pairs": pd.DataFrame([{"Src ENV": "prod", "Dst ENV": "dev", "Connections": 9}])},
                "app": {"top_cross_pairs": pd.DataFrame([{"Src APP": "web", "Dst APP": "db", "Connections": 8}])},
                "role": {"top_cross_pairs": pd.DataFrame([{"Src ROLE": "web", "Dst ROLE": "db", "Connections": 7}])},
                "loc": {"top_cross_pairs": pd.DataFrame([{"Src LOC": "us", "Dst LOC": "eu", "Connections": 6}])},
            },
            "chart_spec": None,
        },
        "mod08": {
            "top_unmanaged_src": pd.DataFrame([{"Src IP": "10.0.0.9", "Connections": 4}]),
            "per_dst_app": pd.DataFrame([{"App": "web", "Connections": 3}]),
            "exposed_ports_merged": pd.DataFrame([{"Port": 445, "Sources": 2}]),
            "src_port_detail": pd.DataFrame([{"Src IP": "10.0.0.9", "Port": 445}]),
            "managed_hosts_targeted_by_unmanaged": pd.DataFrame([{"Dst IP": "10.0.0.2", "Connections": 3}]),
            "top_unmanaged_dst": pd.DataFrame([{"Dst IP": "10.0.0.2", "Connections": 3}]),
            "chart_spec": None,
        },
    }


def _flat(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    return " ".join(str(c) for r in rows for c in r if c is not None)


def test_traffic_xlsx_has_expected_sheets(full_module_results, tmp_path):
    from src.report.report_generator import build_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    expected = {"Executive Summary", "Policy Decisions", "Uncovered Flows",
                "Lateral Movement", "Cross-Label Matrix", "Unmanaged Hosts"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"
    # 舊 Top Talkers sheet 已廢棄（主機層資訊改由 Lateral Movement 的 ip_top_talkers 提供）
    assert "Top Talkers" not in wb.sheetnames


def test_traffic_xlsx_executive_sheet_has_kpis(full_module_results, tmp_path):
    from src.report.report_generator import build_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    flat = _flat(wb["Executive Summary"])
    assert "Total Flows" in flat and "Blocked Flows" in flat, "Executive Summary KPI 標籤缺失"


def test_traffic_xlsx_uncovered_sheet_has_rows(full_module_results, tmp_path):
    from src.report.report_generator import build_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    assert "Uncovered Flows" in wb.sheetnames
    rows = list(wb["Uncovered Flows"].iter_rows(values_only=True))
    assert len(rows) >= 1, "Uncovered Flows sheet has no rows"
    assert "10.0.0.4" in _flat(wb["Uncovered Flows"]), "top_flows 內容應出現"


def test_no_lateral_note_absent_when_any_table_has_data(tmp_path):
    """service_summary 空但下放表有資料時，no-lateral note 不得出現。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    module_results = {
        "mod15": {
            "service_summary": pd.DataFrame(),
            "fan_out_sources": pd.DataFrame(),
            "allowed_lateral_flows": pd.DataFrame(),
            "attack_paths": pd.DataFrame(),
            "ip_top_talkers": pd.DataFrame([{"IP": "10.0.0.1", "Connections": 5}]),
            "ip_top_pairs": pd.DataFrame(),
            "source_risk_scores": pd.DataFrame(),
            "bridge_nodes": pd.DataFrame(),
            "top_reachable_nodes": pd.DataFrame(),
            "app_chains": pd.DataFrame(),
        },
    }
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    flat = _flat(wb["Lateral Movement"])
    assert "10.0.0.1" in flat, "下放表資料應出現於 sheet"
    no_lateral_text = t("rpt_xlsx_no_lateral", lang="en")
    assert no_lateral_text not in flat, f"'{no_lateral_text}' 不應在有資料時出現"


def test_no_lateral_note_present_when_all_tables_empty(tmp_path):
    """所有 lateral 子表皆空時，落回 no-lateral note。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    module_results = {
        "mod15": {
            "service_summary": pd.DataFrame(),
            "ip_top_talkers": pd.DataFrame(),
            "app_chains": pd.DataFrame(),
        },
    }
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    assert t("rpt_xlsx_no_lateral", lang="en") in _flat(wb["Lateral Movement"])


def test_lateral_sheet_has_demoted_table_titles(full_module_results, tmp_path):
    """Lateral sheet 必須含六下放表標題。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    flat = _flat(wb["Lateral Movement"])
    for key in ("rpt_tr_ip_top_talkers", "rpt_tr_ip_top_pairs", "rpt_tr_top_risk_sources",
                "rpt_mod15_bridge_nodes", "rpt_mod15_top_reachable", "rpt_tr_app_chains"):
        title = t(key, lang="en")
        assert title in flat, f"下放表標題 {key}='{title}' 未出現於 lateral sheet"


def test_cross_label_sheet_has_role_loc_demoted_tables(full_module_results, tmp_path):
    """Cross-Label Matrix sheet 必須含 role/loc 下放明細與維度標題。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    sheet_name = t("rpt_xlsx_sheet_cross_label", lang="en")
    assert sheet_name in wb.sheetnames
    flat = _flat(wb[sheet_name])
    assert "web" in flat and "db" in flat  # role 表內容
    assert "us" in flat and "eu" in flat   # loc 表內容
    assert "ROLE" in flat and "LOC" in flat


def test_cross_label_sheet_fallback_note_when_role_loc_empty(tmp_path):
    """role/loc 皆無 top_cross_pairs 時，落回 no-matrix 提示。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    module_results = {
        "mod07": {
            "matrices": {
                "role": {"top_cross_pairs": pd.DataFrame()},
                "loc": {"top_cross_pairs": pd.DataFrame()},
            },
        },
    }
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    sheet_name = t("rpt_xlsx_sheet_cross_label", lang="en")
    assert t("rpt_no_matrix", lang="en") in _flat(wb[sheet_name])


def test_unmanaged_sheet_has_demoted_tables(full_module_results, tmp_path):
    """Unmanaged Hosts sheet 必須含三下放表標題與內容。"""
    from src.report.report_generator import build_traffic_xlsx
    from src.i18n import t

    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(full_module_results, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    sheet_name = t("rpt_xlsx_sheet_unmanaged", lang="en")
    assert sheet_name in wb.sheetnames
    flat = _flat(wb[sheet_name])
    for key in ("rpt_tr_top_unmanaged", "rpt_tr_managed_apps_unmanaged",
                "rpt_tr_exposed_ports_merged", "rpt_tr_src_port_detail",
                "rpt_tr_managed_targeted", "rpt_tr_top_unmanaged_dst"):
        assert t(key, lang="en") in flat, f"下放表標題 {key} 未出現於 unmanaged sheet"


def test_traffic_profile_skips_sec_only_sheets(tmp_path):
    """traffic profile 的 module_results 無 mod15/mod07/mod03 → 該三 sheet 略過且不炸。"""
    from src.report.report_generator import build_traffic_xlsx
    module_results = {
        "mod12": {"kpis": [{"label": "Total Flows", "value": "7"}]},
        "mod02": {"summary": pd.DataFrame([{"Decision": "allowed", "Flows": 5}])},
        "mod08": {"top_unmanaged_src": pd.DataFrame([{"Src IP": "10.0.0.9", "Connections": 4}])},
    }
    out_path = tmp_path / "traffic.xlsx"
    build_traffic_xlsx(module_results, str(out_path), profile="traffic")
    wb = load_workbook(str(out_path))
    names = set(wb.sheetnames)
    assert "Executive Summary" in names
    assert "Policy Decisions" in names
    assert "Unmanaged Hosts" in names
    # 該 profile 不跑 mod03/mod07/mod15 → 對應 sheet 不得出現
    assert "Uncovered Flows" not in names
    assert "Cross-Label Matrix" not in names
    assert "Lateral Movement" not in names


def test_export_path_writes_real_traffic_xlsx(full_module_results, tmp_path):
    """整合測試：ReportGenerator.export(fmt='xlsx') 走完整路徑，讀回應有非空分頁。"""
    from src.report.report_generator import ReportGenerator, ReportResult

    gen = ReportGenerator(config_manager=MagicMock(), api_client=MagicMock())
    result = ReportResult(
        data_source="csv",
        record_count=7,
        date_range=("2026-04-25", "2026-04-25"),
        module_results=full_module_results,
        dataframe=None,
    )
    paths = gen.export(result, fmt="xlsx", output_dir=str(tmp_path),
                       traffic_report_profile="security_risk")
    xlsx_paths = [p for p in paths if p.endswith(".xlsx")]
    assert xlsx_paths, f"export(fmt='xlsx') 應回傳 xlsx path；errors={getattr(gen, 'last_export_errors', None)}"
    wb = load_workbook(xlsx_paths[0])
    assert "Executive Summary" in wb.sheetnames
    non_empty = [n for n in wb.sheetnames
                 if any(any(v is not None for v in row) for row in wb[n].iter_rows(values_only=True))]
    assert non_empty, "至少一分頁需有資料列"
