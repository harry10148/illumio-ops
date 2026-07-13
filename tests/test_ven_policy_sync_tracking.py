"""VEN 報表 policy sync 追蹤（使用者需求 2026-07-13）：

除 offline 追蹤外，偵測「回報正常（online）但 policy sync 異常」的 VEN。
Vendor 事實（REST API 25.4，NotebookLM 查證）：
- security_policy_sync_state 值域僅 applied/syncing/staged，無 error 值；
- 真正錯誤訊號在 agent_health / agent_health_errors（err/warning）；
- security_policy_refresh_at 已棄用 → security_policy_received_at/applied_at；
  received > applied 即「收到未套用」。
"""
from __future__ import annotations

import datetime
from unittest.mock import MagicMock

import pytest

from src.report.ven_status_generator import VenStatusGenerator


def _wl(hostname, hb_hours_ago=0.1, state="applied", errors=None, warnings=None,
        fw_current=True, received=None, applied=None, status="active"):
    now = datetime.datetime.now(datetime.timezone.utc)
    hb = (now - datetime.timedelta(hours=hb_hours_ago)).isoformat()
    st = {
        "status": status,
        "last_heartbeat_on": hb,
        "hours_since_last_heartbeat": hb_hours_ago,
        "security_policy_sync_state": state,
        "fw_config_current": fw_current,
        "security_policy_received_at": received if received is not None else hb,
        "security_policy_applied_at": applied if applied is not None else (received if received is not None else hb),
        "agent_health": [],
        "agent_health_errors": {"errors": errors or [], "warnings": warnings or []},
        "agent_version": "23.2.10",
        "managed_since": "2026-01-01T00:00:00Z",
    }
    return {"hostname": hostname, "interfaces": [], "labels": [],
            "agent": {"status": st}}


def _analyze(workloads):
    cm = MagicMock()
    cm.config = {"settings": {"timezone": "UTC"}}
    gen = VenStatusGenerator(cm, api_client=MagicMock())
    gen._lang = "en"   # generate() 正常路徑會設定；此處直呼內部方法
    df = gen._build_dataframe(workloads)
    return gen._analyze(df)


def _issue_hosts(results):
    df = results["sync_issues"]
    return set(df["Hostname"]) if not df.empty else set()


def test_online_applied_clean_not_flagged():
    r = _analyze([_wl("ok-host")])
    assert _issue_hosts(r) == set()


def test_online_health_error_flagged():
    r = _analyze([_wl("bad-host", errors=["policy_apply_failed"])])
    assert _issue_hosts(r) == {"bad-host"}
    row = r["sync_issues"].iloc[0]
    assert "policy_apply_failed" in row["Sync Issue"]


def test_online_syncing_state_flagged():
    r = _analyze([_wl("sync-host", state="syncing")])
    assert _issue_hosts(r) == {"sync-host"}
    assert "syncing" in r["sync_issues"].iloc[0]["Sync Issue"]


def test_online_received_after_applied_flagged():
    now = datetime.datetime.now(datetime.timezone.utc)
    rcv = now.isoformat()
    app = (now - datetime.timedelta(hours=3)).isoformat()
    r = _analyze([_wl("stale-host", received=rcv, applied=app)])
    assert _issue_hosts(r) == {"stale-host"}


def test_online_fw_config_stale_flagged():
    r = _analyze([_wl("fw-host", fw_current=False)])
    assert _issue_hosts(r) == {"fw-host"}


def test_offline_syncing_not_in_sync_bucket():
    """offline 已有專章追蹤；sync 異常桶只收「回報正常」的 VEN。"""
    r = _analyze([_wl("dead-host", hb_hours_ago=30.0, state="syncing")])
    assert _issue_hosts(r) == set()


def test_kpi_counts_sync_issues():
    r = _analyze([
        _wl("ok"),
        _wl("bad", errors=["policy_apply_failed"]),
        _wl("sync", state="syncing"),
        _wl("dead", hb_hours_ago=30.0, state="syncing"),
    ])
    kpis = {k["i18n_key"]: k["value"] for k in r["kpis"]}
    assert kpis.get("rpt_ven_kpi_sync_issues") == "2"


def test_policy_received_prefers_non_deprecated_field():
    """security_policy_refresh_at 已棄用：received 欄優先取
    security_policy_received_at，缺值才 fallback refresh_at（舊 PCE 相容）。"""
    w = _wl("h1")
    w["agent"]["status"]["security_policy_received_at"] = "2026-07-01T00:00:00Z"
    w["agent"]["status"]["security_policy_refresh_at"] = "2026-06-01T00:00:00Z"
    cm = MagicMock()
    cm.config = {"settings": {"timezone": "UTC"}}
    gen = VenStatusGenerator(cm, api_client=MagicMock())
    df = gen._build_dataframe([w])
    assert df.iloc[0]["policy_received"] == "2026-07-01T00:00:00Z"


def test_policy_sync_badge_treats_applied_as_healthy():
    """PCE 實際值域是 applied/syncing/staged；舊 badge 只認 'synced'，
    導致 applied 被渲染成紅色 unsynced badge。"""
    from src.report.exporters.ven_html_exporter import _policy_sync_badge
    assert "badge-synced" in _policy_sync_badge("applied")
    assert "badge-synced" in _policy_sync_badge("synced")
    assert "badge-staged" in _policy_sync_badge("syncing")
    assert "badge-staged" in _policy_sync_badge("staged")
    assert "badge-unsynced" in _policy_sync_badge("weird")
    assert _policy_sync_badge("") == ""


def test_i18n_strings_registered():
    from src.report.exporters.report_i18n import STRINGS, COL_I18N
    for key in ("rpt_ven_kpi_sync_issues", "rpt_ven_sec_sync_title",
                "rpt_ven_sec_sync_intro", "rpt_ven_nav_sync"):
        assert key in STRINGS, key
        assert len(STRINGS[key]) == 2
    assert "Sync Issue" in COL_I18N and "Policy Applied" in COL_I18N


def test_xlsx_includes_sync_issues_sheet(tmp_path):
    from src.report.ven_status_generator import generate_ven_xlsx
    from openpyxl import load_workbook
    r = _analyze([_wl("bad", errors=["policy_apply_failed"]), _wl("ok")])
    out = tmp_path / "ven.xlsx"
    generate_ven_xlsx(r, str(out), lang="en")
    wb = load_workbook(str(out))
    from src.i18n import t
    assert t("rpt_xlsx_sheet_ven_sync_issues", lang="en") in wb.sheetnames
