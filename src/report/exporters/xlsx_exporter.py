"""openpyxl-based xlsx export for illumio_ops reports.

One sheet per analysis module + a Summary sheet. Header row frozen,
alternate-row banding for readability, red fill on 'blocked' / 'deny'
rows. chart_spec (if present) rendered as matplotlib PNG and embedded.
"""
from __future__ import annotations

import io
import pandas as pd
from loguru import logger
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.i18n import t
from src.report.exporters.chart_renderer import render_matplotlib_png

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="375379")
_ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")
_ALERT_TOKENS = ("blocked", "deny", "violat", "critical", "red_flag")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _neutralize(val: Any) -> Any:
    """Defuse spreadsheet/CSV formula injection. A str that Excel would treat as a
    formula (leading = + - @) is prefixed with a single quote so openpyxl stores it
    as text rather than a live formula. Non-str values pass through unchanged."""
    if isinstance(val, str) and val[:1] in _FORMULA_PREFIXES:
        return "'" + val
    return val


def _safe_sheet_name(name: str) -> str:
    # openpyxl sheet names capped at 31 chars and cannot contain :\/?*[]
    return "".join(c for c in name if c not in r"\/:?*[]")[:31] or "Sheet"


def _clean_nan(val: Any) -> Any:
    """NaN/None 等缺值一律寫為空字串；其餘原樣返回。"""
    if pd.isna(val):
        return ""
    return val


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    """依欄名與資料長度粗略估算欄寬（上限 60）。"""
    for col_idx, col in enumerate(df.columns, 1):
        values = [_clean_nan(v) for v in df[col].tolist()]
        max_len = max(len(str(col)), max((len(str(v)) for v in values), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _embed_chart(ws, chart_spec: dict[str, Any], safe_name: str, anchor_row: int) -> None:
    try:
        png = render_matplotlib_png(chart_spec)
        img = XLImage(io.BytesIO(png))
        img.anchor = f"A{anchor_row}"
        ws.add_image(img)
    except Exception as exc:
        logger.warning("Failed to render chart for {}: {}", safe_name, exc)


def append_df_rows(ws, df: pd.DataFrame, *, header: bool = True) -> None:
    """DataFrame 逐列寫入現有 sheet：header 列 + 資料列；每格經 _neutralize；
    列文字含 _ALERT_TOKENS 者套 _ALERT_FILL；NaN 寫空字串。"""
    columns = [str(c) for c in df.columns]
    if header:
        ws.append([_neutralize(c) for c in columns])
        for cell in ws[ws.max_row]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    for record in df.to_dict("records"):
        row_vals = [_clean_nan(record.get(c)) for c in columns]
        row_text = " ".join(str(v).lower() for v in row_vals)
        is_alert = any(tok in row_text for tok in _ALERT_TOKENS)
        ws.append([_neutralize(v) for v in row_vals])
        if is_alert:
            for cell in ws[ws.max_row]:
                cell.fill = _ALERT_FILL


def add_df_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame | None,
    *,
    empty_note: str | None = None,
    chart_spec: dict[str, Any] | None = None,
    lang: str = "en",
) -> None:
    """建立新 sheet 並寫入 DataFrame；df 為空/None 時寫入單行提示列；chart_spec 非 None 時嵌入圖表。"""
    safe_name = _safe_sheet_name(title)
    ws = wb.create_sheet(title=safe_name)

    if df is not None and not df.empty:
        append_df_rows(ws, df)
        ws.freeze_panes = "A2"
        _autosize_columns(ws, df)
        anchor_row = 1 + len(df) + 3
    else:
        ws.append([t("rpt_xlsx_col_note", lang=lang), empty_note or t("rpt_xlsx_no_data", lang=lang)])
        anchor_row = ws.max_row + 2

    if chart_spec:
        _embed_chart(ws, chart_spec, safe_name, anchor_row)


def add_stacked_tables_sheet(
    wb: Workbook,
    title: str,
    tables: list[tuple[str, pd.DataFrame]],
    *,
    empty_note: str | None = None,
    lang: str = "en",
) -> None:
    """一張 sheet 內堆疊多組 (subtitle, df)：逐表以空列 + 粗體標題列 + append_df_rows 呈現；
    全部為空時寫入單行 empty_note 提示。"""
    safe_name = _safe_sheet_name(title)
    ws = wb.create_sheet(title=safe_name)
    non_empty = [(subtitle, df) for subtitle, df in tables if df is not None and not df.empty]

    if not non_empty:
        ws.append([t("rpt_xlsx_col_note", lang=lang), empty_note or t("rpt_xlsx_no_data", lang=lang)])
        return

    for idx, (subtitle, df) in enumerate(non_empty):
        if idx > 0:
            ws.append([])  # 表格之間空列分隔
        ws.append([_neutralize(str(subtitle))])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        append_df_rows(ws, df)


def _write_module_sheet(wb: Workbook, name: str, module_data: dict[str, Any]) -> None:
    safe_name = _safe_sheet_name(name)
    ws = wb.create_sheet(title=safe_name)

    start_row = 1
    summary = module_data.get("summary")
    if summary:
        ws.cell(row=1, column=1, value=str(summary)).font = Font(italic=True)
        ws.append([])  # 跳過第 2 列作為視覺間距，沿用舊邏輯 row += 2
        start_row = 3

    table = module_data.get("table") or []
    if table:
        headers = list(table[0].keys())
        df = pd.DataFrame(table, columns=headers)
        append_df_rows(ws, df)
        header_row = start_row
        ws.freeze_panes = f"A{header_row + 1}"
        _autosize_columns(ws, df)
        anchor_row = header_row + len(table) + 3
    else:
        anchor_row = start_row + 2

    chart_spec = module_data.get("chart_spec")
    if chart_spec:
        _embed_chart(ws, chart_spec, safe_name, anchor_row)

def export_xlsx(report_result: dict[str, Any], output_path: str) -> None:
    """Export a ReportResult-shaped dict to an .xlsx file."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    meta = report_result.get("metadata", {})
    summary_ws["A1"] = meta.get("title", "Report")
    summary_ws["A1"].font = Font(size=18, bold=True)
    summary_ws["A2"] = f"Generated: {meta.get('generated_at', '')}"
    if meta.get("start_date"):
        summary_ws["A3"] = f"Period: {meta.get('start_date')} \u2192 {meta.get('end_date', '')}"
    summary_ws["A4"] = f"Records: {report_result.get('record_count', 0)}"
    summary_ws.freeze_panes = "A2"

    for mod_name, mod_data in (report_result.get("module_results") or {}).items():
        _write_module_sheet(wb, mod_name, mod_data)

    wb.save(output_path)
    logger.info("xlsx report written to {}", output_path)
